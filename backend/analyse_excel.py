import openpyxl
import polars as pl
import io

FILE = r'D:\Project\Stock Indicator Testing Tool\Signals-By-PREMIUM-NIFTY-&-BANKNIFTY.xlsx'

# --- Method 1: openpyxl to see RAW cell values (what Excel actually stores) ---
print("=== RAW CELL VALUES (openpyxl) ===")
wb = openpyxl.load_workbook(FILE, data_only=True)
ws = wb.active
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
print("Headers:", headers)
print()

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=False)):
    for j, cell in enumerate(row):
        if cell.value is not None:
            print(f"  Row{i+1} Col[{headers[j] if j < len(headers) else j}]: value={repr(cell.value)}  type={type(cell.value).__name__}  numFmt={cell.number_format}")
    print()

# --- Method 2: Polars raw string read ---
print("\n=== POLARS RAW STRING READ ===")
df = pl.read_excel(FILE, infer_schema_length=0)
print("Shape:", df.shape)
print("Columns:", df.columns)
print("\nFirst 3 rows:")
for r in df.head(3).to_dicts():
    print(r)
