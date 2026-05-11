"""
services/results.py — Backtest result retrieval and export engine.

REUSABLE PATTERNS:
- StreamingResponse avoids buffering entire file in RAM; safe for large exports.
- CSV export divides stored integers by 100 to restore decimal prices.
- PDF uses reportlab (no headless browser / wkhtmltopdf dependency needed).
- Filename convention: Backtest_{Stock}_{Indicator}_{Timestamp}.{ext}
- GET /api/options-data provides OHLC data for the per-trade Verify chart.

KNOWN BUG AVOIDED: Do not use Response(content=bytes) for large files —
it buffers everything in memory. Use StreamingResponse(io.BytesIO(...)) instead.
"""

import io
import csv
import json
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select

from database import get_session
from models import OptionsData, ValidationReport

router = APIRouter(prefix="/api", tags=["results"])


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/results — List all saved backtest reports (summary only, fast)
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/results")
def list_results(session: Session = Depends(get_session)):
    """
    Returns all saved ValidationReport records for the Results History panel.
    Uses the indexed summary columns so JSON parsing of config/trades is avoided.
    """
    stmt = select(ValidationReport).order_by(ValidationReport.testDate.desc())
    reports = session.exec(stmt).all()
    result = []
    for r in reports:
        # Fallback to config JSON if summary columns are NULL (legacy rows)
        config = {}
        try:
            config = json.loads(r.config)
        except Exception:
            pass
        result.append({
            "id":            r.id,
            "testDate":      r.testDate.isoformat(),
            "stock":         r.stock or config.get("stock", "—"),
            "indicatorName": r.indicatorName or config.get("indicatorName", "—"),
            "timeframe":     config.get("timeframe", "—"),
            "winRate":       round(r.winRate, 2),
            "totalProfit":   round(r.totalProfit / 100.0, 2),
            "totalTrades":   r.totalTrades or 0,
        })
    return result


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/results/export-excel — Stream .xlsx report with user-defined columns
# KNOWN BUG AVOIDED: This MUST be defined before /api/results/{report_id} to
# prevent FastAPI capturing "export-excel" as a report_id path param.
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/results/export-excel")
def export_excel(
    resultId: str = Query(...),
    session:  Session = Depends(get_session)
):
    """
    Generates and streams an .xlsx file for the specified report.
    Columns: Final Entry Script, Option Type, Expiry, Entry Time, Entry AT (Value),
             Buy Amount, Exit Time, Exit At (Value), Sell Amount, PnL Points,
             PnL Amount, PnL Percentage, Highest High Percentage, Lowest Low Percentage.
    REUSABLE: This openpyxl pattern works for any list-of-dicts → styled Excel export.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    report = session.get(ValidationReport, resultId)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found.")

    trades = []
    try:
        trades = json.loads(report.trades)
    except Exception:
        pass

    stock     = report.stock or "UNKNOWN"
    indicator = report.indicatorName or "UNKNOWN"
    timestamp = report.testDate.strftime("%Y%m%d_%H%M")

    # --- Column definition: (header label, trade dict key) ---
    # REUSABLE: Changing this list updates both the sheet and header row automatically.
    # Columns 13-16 are split: Highest (raw), High % (ratio), Lowest (raw), Lowest % (ratio)
    COLUMNS = [
        ("Final Entry Script",  "script"),
        ("Option Type",         "optionType"),
        ("Expiry",              "expiry"),
        ("Entry Time",          "entryTime"),
        ("Entry AT (Value)",    "entryPrice"),
        ("Buy Amount",          "tradeValue"),
        ("Exit Time",           "exitTime"),
        ("Exit At (Value)",     "exitPrice"),
        ("Sell Amount",         "sellAmount"),
        ("PnL Points",          "points"),
        ("PnL Amount",          "profit"),
        ("PnL Percentage",      "pnlPct"),
        ("Highest",             "highestHigh"),      # Max High value between signals
        ("High Percentage",     "highestHighPct"),   # highestHigh / entryPrice
        ("Lowest",              "lowestLow"),         # Min Low value between signals
        ("Lowest Percentage",   "lowestLowPct"),     # lowestLow / entryPrice
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Header style: yellow background, bold black text (matches user's design)
    HEADER_FILL  = PatternFill("solid", fgColor="FFFF00")
    HEADER_FONT  = Font(bold=True, color="000000")
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN_BORDER  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # Write report title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    title_cell = ws.cell(row=1, column=1, value=f"Report — {stock} | {indicator} | {timestamp}")
    title_cell.font = Font(bold=True, size=12, color="000000")
    title_cell.fill = PatternFill("solid", fgColor="4EDEA3")
    title_cell.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 22

    # Write header row
    for col_idx, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 36

    # Write data rows
    PNL_KEYS = {"points", "profit", "pnlPct"}  # Keys to colour green/red based on sign
    for row_idx, trade in enumerate(trades, start=3):
        for col_idx, (_, key) in enumerate(COLUMNS, start=1):
            val = trade.get(key, "")
            if val is None:
                val = ""
            # Format datetime strings for readability
            if isinstance(val, str) and "T" in val:
                val = val.replace("T", " ")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
            # Colour PnL columns green/red based on sign
            if key in PNL_KEYS and val != "":
                try:
                    cell.font = Font(color="006100" if float(val) >= 0 else "9C0006")
                except (ValueError, TypeError):
                    pass

    # Auto-fit column widths
    for col_idx, (label, _) in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(len(label) + 4, 18)

    # Stream the file
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"Report_{stock}_{indicator}_{timestamp}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/results/export — Stream CSV or PDF of a specific backtest report
#
# KNOWN BUG AVOIDED: This route MUST be defined BEFORE /api/results/{report_id}.
# FastAPI matches routes top-to-bottom. If {report_id} is first, the literal
# word 'export' gets captured as a report_id value → 404 "Report 'export' not found".
# Rule: always put static path segments before dynamic {param} routes.
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/results/export")
def export_results(
    resultId: str = Query(...),
    format:   str = Query("csv"),
    session:  Session = Depends(get_session)
):
    """
    Streams a file download of the specified report.
    - format=csv  → Full trade log spreadsheet (prices as decimals).
    - format=pdf  → Summary report with KPIs and trade table via reportlab.
    Filename: Backtest_{Stock}_{Indicator}_{Timestamp}.{ext}
    """
    report = session.get(ValidationReport, resultId)
    if not report:
        raise HTTPException(status_code=404, detail="Report not found.")

    config = {}
    try:
        config = json.loads(report.config)
    except Exception:
        pass

    trades = []
    try:
        trades = json.loads(report.trades)
    except Exception:
        pass

    stock     = report.stock or config.get("stock", "UNKNOWN")
    indicator = report.indicatorName or config.get("indicatorName", "UNKNOWN")
    timestamp = report.testDate.strftime("%Y%m%d_%H%M")

    if format == "csv":
        return _stream_csv(trades, stock, indicator, timestamp, report)
    elif format == "pdf":
        return _stream_pdf(trades, stock, indicator, timestamp, report, config)
    else:
        raise HTTPException(status_code=400, detail="Invalid format. Use 'csv' or 'pdf'.")


# ─────────────────────────────────────────────────────────────────────────────
# GET /api/results/{report_id} — Full details of a single report
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/results/{report_id}")
def get_result(report_id: str, session: Session = Depends(get_session)):
    """Returns full details including all trade rows for a specific report."""
    report = session.get(ValidationReport, report_id)
    if not report:
        raise HTTPException(status_code=404, detail=f"Report '{report_id}' not found.")

    config = {}
    try:
        config = json.loads(report.config)
    except Exception:
        pass

    trades = []
    try:
        trades = json.loads(report.trades)
    except Exception:
        pass

    return {
        "reportId":     report.id,
        "testDate":     report.testDate.isoformat(),
        "stock":        report.stock or config.get("stock", "—"),
        "indicatorName": report.indicatorName or config.get("indicatorName", "—"),
        "totalProfit":  round(report.totalProfit / 100.0, 2),
        "winRate":      round(report.winRate, 2),
        "totalTrades":  report.totalTrades or len(trades),
        "maxDrawdown":  report.maxDrawdown or 0.0,
        "profitFactor": report.profitFactor or 0.0,
        "avgTrade":     report.avgTrade or 0.0,
        "trades":       trades,
        "config":       config,
    }




# ─────────────────────────────────────────────────────────────────────────────
# GET /api/options-data — OHLC candle data for Trade Verify chart
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/options-data")
def get_options_data(
    script:  str = Query(..., description="Exact script name, e.g. NIFTY24OCT24000PE"),
    from_dt: str = Query(..., description="ISO datetime, e.g. 2026-03-11T06:29:00"),
    to_dt:   str = Query(..., description="ISO datetime, e.g. 2026-03-11T09:00:00"),
    session: Session = Depends(get_session)
):
    """
    Returns OHLCV rows for a specific option script in a time window.
    Used by the frontend Trade Verify modal to draw the price chart.
    """
    try:
        from_datetime = datetime.fromisoformat(from_dt)
        to_datetime   = datetime.fromisoformat(to_dt)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid datetime format. Use ISO 8601.")

    stmt = (
        select(OptionsData)
        .where(
            OptionsData.script == script,
            OptionsData.dateTime >= from_datetime,
            OptionsData.dateTime <= to_datetime,
        )
        .order_by(OptionsData.dateTime)
    )
    records = session.exec(stmt).all()

    return [
        {
            "dateTime": r.dateTime.isoformat(),
            "open":     r.open,
            "high":     r.high,
            "low":      r.low,
            "close":    r.close,
            "volume":   r.volume,
        }
        for r in records
    ]


# ─────────────────────────────────────────────────────────────────────────────
# Private: CSV generation helper
# ─────────────────────────────────────────────────────────────────────────────
def _stream_csv(trades, stock, indicator, timestamp, report):
    """Generates and streams a UTF-8 CSV of the full trade log."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Summary header rows
    writer.writerow(["Backtest Report", f"{stock} | {indicator}", f"Generated: {timestamp}"])
    writer.writerow([])
    writer.writerow(["Net P&L (pts)", "Win Rate", "Total Trades", "Max Drawdown", "Profit Factor", "Avg Trade"])
    writer.writerow([
        round(report.totalProfit / 100.0, 2),
        f"{round(report.winRate, 2)}%",
        report.totalTrades or len(trades),
        f"{report.maxDrawdown or 0.0:.2f}%",
        f"{report.profitFactor or 0.0:.2f}x",
        f"{report.avgTrade or 0.0:.2f}",
    ])
    writer.writerow([])

    # Trade log column headers
    writer.writerow([
        "#", "Script", "ATM Proof", "Entry Time", "Entry Type",
        "Exit Reason", "Quantity", "Trade Value", "Entry Price", "Exit Time", "Exit Price",
        "Duration", "Points", "Net P&L", "P&L %",
    ])

    for trade in trades:
        writer.writerow([
            trade.get("tradeId",       ""),
            trade.get("script",        ""),
            trade.get("atmProof",      ""),
            trade.get("entryTime",     ""),
            trade.get("entryType",     ""),
            trade.get("exitReason",    trade.get("executionNote", "")),
            trade.get("quantity",      1),
            trade.get("tradeValue",    0),
            trade.get("entryPrice",    0),   # Already decimal from validator
            trade.get("exitTime",      ""),
            trade.get("exitPrice",     0),
            trade.get("duration",      ""),
            trade.get("points",        0),
            trade.get("profit",        0),
            f"{trade.get('pnlPct', 0)}%",
        ])

    csv_bytes = output.getvalue().encode("utf-8-sig")  # utf-8-sig for Excel compatibility
    filename  = f"Backtest_{stock}_{indicator}_{timestamp}.csv"
    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# Private: PDF generation helper (reportlab)
# ─────────────────────────────────────────────────────────────────────────────
def _stream_pdf(trades, stock, indicator, timestamp, report, config):
    """Generates and streams a PDF summary report using reportlab."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, HRFlowable,
        )
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="reportlab not installed. Run: pip install reportlab"
        )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)

    # Theme palette (matches UI dark theme)
    PRIMARY  = colors.HexColor("#4edea3")
    DARK_BG  = colors.HexColor("#0b1326")
    SURFACE  = colors.HexColor("#171f33")
    MUTED    = colors.HexColor("#909097")
    RED_COL  = colors.HexColor("#ffb4ab")
    WHITE    = colors.HexColor("#dae2fd")

    styles   = getSampleStyleSheet()
    title_st = ParagraphStyle("T", parent=styles["Heading1"], textColor=PRIMARY,  fontSize=20, spaceAfter=4)
    sub_st   = ParagraphStyle("S", parent=styles["Normal"],   textColor=MUTED,    fontSize=9,  spaceAfter=10)
    label_st = ParagraphStyle("L", parent=styles["Normal"],   textColor=WHITE,    fontSize=8)

    story = []

    # Title
    story.append(Paragraph("Backtest Performance Report", title_st))
    story.append(Paragraph(
        f"Strategy: {stock} | Indicator: {indicator} ({config.get('timeframe', '1m')}) | Generated: {timestamp.replace('_', ' at ')}",
        sub_st
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=PRIMARY, spaceAfter=12))

    # KPI summary table
    pnl = round(report.totalProfit / 100.0, 2)
    kpi_data = [
        ["Net P&L", "Win Rate", "Total Trades", "Max Drawdown", "Profit Factor", "Avg Trade"],
        [
            f"{pnl:+.2f} pts",
            f"{round(report.winRate, 2):.1f}%",
            str(report.totalTrades or len(trades)),
            f"{report.maxDrawdown or 0.0:.2f}%",
            f"{report.profitFactor or 0.0:.2f}x",
            f"{report.avgTrade or 0.0:.2f}",
        ],
    ]
    kpi_table = Table(kpi_data, colWidths=[3 * cm] * 6)
    kpi_table.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), SURFACE),
        ("BACKGROUND",  (0, 1), (-1, 1), DARK_BG),
        ("TEXTCOLOR",   (0, 0), (-1, 0), MUTED),
        ("TEXTCOLOR",   (0, 1), (-1, 1), PRIMARY if pnl >= 0 else RED_COL),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("ALIGN",       (0, 0), (-1, -1), "CENTER"),
        ("BOX",         (0, 0), (-1, -1), 0.5, PRIMARY),
        ("INNERGRID",   (0, 0), (-1, -1), 0.25, MUTED),
        ("TOPPADDING",  (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 0.5 * cm))

    # Trade log header
    story.append(Paragraph("Trade Log", ParagraphStyle("TH", parent=styles["Heading2"], textColor=WHITE, fontSize=12, spaceBefore=10, spaceAfter=6)))

    if trades:
        log_headers = ["#", "Script", "Entry Time", "Entry Px", "Exit Time", "Exit Px", "Pts", "P&L%"]
        log_data    = [log_headers]
        for t in trades:
            pts = t.get("points", t.get("profit", 0))
            log_data.append([
                str(t.get("tradeId", "")),
                str(t.get("script",  ""))[:20],
                str(t.get("entryTime", ""))[:16].replace("T", " "),
                f"{t.get('entryPrice', 0):.2f}",
                str(t.get("exitTime",  ""))[:16].replace("T", " "),
                f"{t.get('exitPrice',  0):.2f}",
                f"{pts:+.2f}",
                f"{t.get('pnlPct', 0):.1f}%",
            ])

        col_widths = [1 * cm, 4 * cm, 3.5 * cm, 2 * cm, 3.5 * cm, 2 * cm, 1.8 * cm, 1.8 * cm]
        log_table  = Table(log_data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            ("BACKGROUND",    (0, 0), (-1, 0), SURFACE),
            ("TEXTCOLOR",     (0, 0), (-1, 0), PRIMARY),
            ("FONTSIZE",      (0, 0), (-1, -1), 7),
            ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
            ("BOX",           (0, 0), (-1, -1), 0.5, MUTED),
            ("INNERGRID",     (0, 0), (-1, -1), 0.2, SURFACE),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("BACKGROUND",    (0, 1), (-1, -1), DARK_BG),
            ("TEXTCOLOR",     (0, 1), (-1, -1), WHITE),
        ]
        # Colour P&L column conditionally
        for i, t in enumerate(trades, start=1):
            pts = t.get("points", t.get("profit", 0))
            col_c = PRIMARY if pts >= 0 else RED_COL
            style_cmds.append(("TEXTCOLOR", (6, i), (7, i), col_c))

        log_table.setStyle(TableStyle(style_cmds))
        story.append(log_table)
    else:
        story.append(Paragraph("No trades recorded.", label_st))

    doc.build(story)
    buffer.seek(0)
    filename = f"Backtest_{stock}_{indicator}_{timestamp}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
