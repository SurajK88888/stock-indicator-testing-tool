"""
signal_validator.py — Signal Backtesting Engine.

COMPLETELY ISOLATED from the Indicator Validator (validator.py).
No shared state, no shared logic, no shared models.

Business Rules (from Signal-Validation.docx + Book1.xlsx + user confirmation):
- Entry Rule:
    At Signal / Next Candle: check if entry_price is within candle OHLC.
      - If YES → enter at signal entry_price
      - If candle ABOVE entry_price → Jump Rule (Drop Signal or Trade at selected price)
      - If candle BELOW entry_price → enter at candle close
- SL Hit: low <= SL → exit at SL value (not candle low)
- Target Hit: high >= Target → exit at Target value (not candle high)
- Trailing SL: after T1 hit → SL shifts to entry. After T2 → SL shifts to T1. etc.
- Split Rule: Lot-1 targets T1, Lot-2 targets highest available T with trailing SL.
- Single Trade: all lots together, target = highest available T, trailing SL.
- Global Rule: ALL open positions force-closed on Expiry Date at close price of last candle.
- Intraday: if still open, close at last candle close of same day.
- BTST: if still open, close at last candle close of next day.

REUSABLE PATTERNS:
- Background job dict (_sv_jobs) mirrors validator.py job pattern for consistency.
- _get_entry_price() and _check_jump() centralise entry logic reusable across trade modes.
"""

import json
import uuid
from typing import Optional
from datetime import datetime, timedelta, date

from fastapi import APIRouter, BackgroundTasks, Depends
from pydantic import BaseModel
from sqlmodel import Session, select

from database import get_session, sqlite_url
from models import SignalData, OptionsData, SignalValidationReport

router = APIRouter(prefix="/api/signal-validate", tags=["signal_validator"])

# In-memory job store — same pattern as validator.py
# REUSABLE: This dict-based job store works for low concurrency local apps.
# Replace with Redis or a DB-backed queue for production scale.
_sv_jobs: dict = {}


# ---------------------------------------------------------------------------
# Request Schema
# ---------------------------------------------------------------------------
class SignalValidateRequest(BaseModel):
    signal_provider: str
    stock: str                              # "ALL" or specific stock name
    start_date: Optional[str] = None        # "YYYY-MM-DD"
    end_date: Optional[str] = None          # "YYYY-MM-DD"
    entry_time: str = "At Signal"           # "At Signal" | "Next Candle"
    on_jump: str = "Drop Signal"            # "Drop Signal" | "Trade"
    entry_value_if_jump: str = "Close"      # "Open"|"High"|"Low"|"Close"|"Open-Close Average"|"High-Low Average"
    trade_on: str = "Calls"                 # "Calls" | "Puts" | "Both"
    trade_amount: int = 1                   # Number of lots (must be even for Split Rule)
    lot_split_rule: str = "Single Trade"    # "Split Rule" | "Single Trade"
    closing_on_trade_type: str = "Ignore last Entry"   # "Ignore last Entry" | "Take next day beyond End Date" | "Close at EOD Values"
    position_open_on_end_date: str = "Ignore last Entry"  # same options


# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def _price_from_candle(candle: OptionsData, price_point: str) -> float:
    """
    Returns the price from a candle row based on the selected price point.
    REUSABLE: Same logic as validator.py _get_price_by_point, kept isolated here.
    """
    if price_point == "Open-Close Average":
        return round((candle.open + candle.close) / 2.0, 4)
    if price_point == "High-Low Average":
        return round((candle.high + candle.low) / 2.0, 4)
    mapping = {"Open": candle.open, "High": candle.high, "Low": candle.low, "Close": candle.close}
    return round(mapping.get(price_point, candle.close), 4)


def _get_entry_price(candle: OptionsData, signal_entry_price: float, on_jump: str, entry_value_if_jump: str):
    """
    Determines final entry price from candle and signal entry price.

    Entry Rule (confirmed by user):
    1. If signal_entry_price is within candle OHLC (low <= price <= high) → use signal_entry_price
    2. If candle is ABOVE entry price (candle low > signal_entry_price) → Jump Rule:
       - "Drop Signal" → return None (skip)
       - "Trade" → use the selected price point
    3. If candle is BELOW entry price (candle high < signal_entry_price) → enter at candle close

    Returns: (entry_price: float | None, jump_applied: bool)
    None means signal should be dropped.
    """
    lo, hi = candle.low, candle.high

    # Case 1: entry_price is within OHLC range
    if lo <= signal_entry_price <= hi:
        return signal_entry_price, False

    # Case 2: candle is ABOVE entry price (candle low > signal entry price)
    if lo > signal_entry_price:
        if on_jump == "Drop Signal":
            return None, True   # Drop this signal
        else:
            return _price_from_candle(candle, entry_value_if_jump), True

    # Case 3: candle is BELOW entry price (candle high < signal entry price)
    return candle.close, False


def _highest_target(signal: SignalData) -> Optional[float]:
    """
    Returns the highest non-null target value from T1–T10.
    REUSABLE: Used to find the final target in Single Trade mode.
    """
    targets = [
        signal.target_1, signal.target_2, signal.target_3, signal.target_4,
        signal.target_5, signal.target_6, signal.target_7, signal.target_8,
        signal.target_9, signal.target_10
    ]
    valid = [t for t in targets if t is not None and t > 0]
    return max(valid) if valid else None


def _target_list(signal: SignalData) -> list:
    """
    Returns ordered list of (index, value) for all non-null targets T1–T10.
    Index is 1-based (1=T1 ... 10=T10).
    """
    raw = [
        signal.target_1, signal.target_2, signal.target_3, signal.target_4,
        signal.target_5, signal.target_6, signal.target_7, signal.target_8,
        signal.target_9, signal.target_10
    ]
    return [(i + 1, v) for i, v in enumerate(raw) if v is not None and v > 0]


def _format_dt(dt: Optional[datetime]) -> Optional[str]:
    """Format datetime as 'YYYY-MM-DD HH:MM:SS' string, or None."""
    return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else None


def _format_time(dt: Optional[datetime]) -> Optional[str]:
    """Format datetime as 'HH:MM' string, or '-'."""
    return dt.strftime("%H:%M") if dt else "-"


def _last_candle_of_day(candles_by_date: dict, target_date: date) -> Optional[OptionsData]:
    """Returns the last candle of a given date from a pre-grouped dict."""
    day_candles = candles_by_date.get(target_date, [])
    return day_candles[-1] if day_candles else None


# ---------------------------------------------------------------------------
# Core Trade Engine
# ---------------------------------------------------------------------------

def _run_trade(
    signal: SignalData,
    candles: list,          # OptionsData rows sorted by dateTime for this script+expiry window
    candles_by_date: dict,  # date → [OptionsData] for EOD lookup
    req: SignalValidateRequest,
    serial_num: int
) -> tuple:
    """
    Runs backtesting for a single signal against ordered candle data.
    Returns (summary_row: dict, validation_row: dict) or (None, None) if skipped.

    DESIGN:
    - This function is pure logic — no DB access, no side effects.
    - All state is local to this call — safe for concurrent/parallel use in future.
    - Summary row = Report 1 format (docx). Validation row = Report 2 format.
    REUSABLE: Extend this function to support more trade rule types.
    """
    if not candles:
        return None, None

    # --- Determine entry candle ---
    if req.entry_time == "Next Candle":
        # Skip first candle (signal candle), use next one
        entry_candles = candles[1:]
        if not entry_candles:
            return None, None
        entry_candle = entry_candles[0]
    else:
        entry_candle = candles[0]

    # --- Entry Price ---
    final_entry_price, jump_applied = _get_entry_price(
        entry_candle, signal.entry_price, req.on_jump, req.entry_value_if_jump
    )
    if final_entry_price is None:
        return None, None   # Signal dropped due to jump rule

    entry_time = entry_candle.dateTime
    entry_type_label = req.entry_time + (" (Jump)" if jump_applied else "")

    # --- Lot size from candle data (dynamic, not hardcoded) ---
    lot_size = entry_candle.lot_size if entry_candle.lot_size and entry_candle.lot_size > 0 else 1
    total_lots = req.trade_amount
    qty = total_lots * lot_size
    trade_value_entry = round(qty * final_entry_price, 2)

    # --- Prepare targets list ---
    all_targets = _target_list(signal)   # [(idx, value), ...]
    sl_value = signal.sl
    signal_date = entry_candle.dateTime.date()
    expiry_date = None
    if signal.expiry:
        try:
            expiry_date = date.fromisoformat(signal.expiry)
        except ValueError:
            expiry_date = None

    # --- Target hit time tracking (for Validation Report) ---
    # Each entry: None until hit, then set to datetime
    target_hit_times: dict = {idx: None for idx, _ in all_targets}
    sl_hit_time: Optional[datetime] = None

    # ====================================================================
    # SPLIT RULE MODE
    # ====================================================================
    if req.lot_split_rule == "Split Rule":
        half_lots = max(1, total_lots // 2)
        half_qty = half_lots * lot_size

        # -- Lot-1: target = T1, SL = signal SL --
        l1_exit_time: Optional[datetime] = None
        l1_exit_price: Optional[float] = None

        # -- Lot-2: target = highest T, trailing SL --
        l2_exit_time: Optional[datetime] = None
        l2_exit_price: Optional[float] = None

        t1_value = all_targets[0][1] if all_targets else None
        highest_t_value = _highest_target(signal)

        l1_done = t1_value is None   # if no T1, treat Lot-1 as done immediately
        l2_done = highest_t_value is None

        current_sl = sl_value   # Lot-2 trailing SL starts at signal SL
        last_hit_target_value = None  # for trailing SL reference

        # Candles to iterate — start after entry candle
        iter_start = 1 if req.entry_time == "At Signal" else 2
        trade_candles = candles[iter_start:]

        for candle in trade_candles:
            c_date = candle.dateTime.date()
            trade_type = (signal.trade_type or "Intraday").strip()

            # Global Rule: force close on expiry date at close of last candle of expiry day
            if expiry_date and c_date > expiry_date:
                last_expiry_candle = _last_candle_of_day(candles_by_date, expiry_date)
                close_price = last_expiry_candle.close if last_expiry_candle else final_entry_price
                close_time = last_expiry_candle.dateTime if last_expiry_candle else candle.dateTime
                if not l1_done:
                    l1_exit_time, l1_exit_price = close_time, close_price
                    l1_done = True
                if not l2_done:
                    l2_exit_time, l2_exit_price = close_time, close_price
                    l2_done = True
                break

            # Intraday: close at EOD if configured
            if trade_type == "Intraday" and req.closing_on_trade_type == "Close at EOD Values":
                if c_date > signal_date:
                    last_c = _last_candle_of_day(candles_by_date, signal_date)
                    cp = last_c.close if last_c else final_entry_price
                    ct = last_c.dateTime if last_c else candle.dateTime
                    if not l1_done:
                        l1_exit_time, l1_exit_price = ct, cp
                        l1_done = True
                    if not l2_done:
                        l2_exit_time, l2_exit_price = ct, cp
                        l2_done = True
                    break

            # BTST: close at end of next day
            if trade_type == "BTST" and req.closing_on_trade_type == "Close at EOD Values":
                next_day = signal_date + timedelta(days=1)
                if c_date > next_day:
                    last_c = _last_candle_of_day(candles_by_date, next_day)
                    cp = last_c.close if last_c else final_entry_price
                    ct = last_c.dateTime if last_c else candle.dateTime
                    if not l1_done:
                        l1_exit_time, l1_exit_price = ct, cp
                        l1_done = True
                    if not l2_done:
                        l2_exit_time, l2_exit_price = ct, cp
                        l2_done = True
                    break

            # Record target hit times (all targets, for Validation Report)
            for idx, t_val in all_targets:
                if target_hit_times[idx] is None and candle.high >= t_val:
                    target_hit_times[idx] = candle.dateTime

            # SL check (both lots share initial SL; Lot-2 uses trailing)
            sl_hit = candle.low <= sl_value
            if sl_hit and sl_hit_time is None:
                sl_hit_time = candle.dateTime

            # Lot-1 logic
            if not l1_done:
                if candle.low <= sl_value:
                    l1_exit_time, l1_exit_price = candle.dateTime, sl_value
                    l1_done = True
                elif t1_value and candle.high >= t1_value:
                    l1_exit_time, l1_exit_price = candle.dateTime, t1_value
                    l1_done = True

            # Lot-2 logic (trailing SL)
            if not l2_done:
                # Update trailing SL after each target crossed
                new_sl = current_sl
                for idx, t_val in all_targets:
                    if candle.high >= t_val:
                        # SL trails to previous level
                        if idx == 1:
                            new_sl = max(new_sl, final_entry_price)
                        else:
                            prev_val = all_targets[idx - 2][1]  # idx is 1-based, list is 0-based
                            new_sl = max(new_sl, prev_val)
                        last_hit_target_value = t_val
                current_sl = new_sl

                # Check trailing SL
                if candle.low <= current_sl:
                    l2_exit_time, l2_exit_price = candle.dateTime, current_sl
                    l2_done = True
                # Check highest target
                elif highest_t_value and candle.high >= highest_t_value:
                    l2_exit_time, l2_exit_price = candle.dateTime, highest_t_value
                    l2_done = True

            if l1_done and l2_done:
                break

        # Fallback: if still open (no close trigger found)
        if not l1_done:
            last_c = candles[-1]
            l1_exit_time, l1_exit_price = last_c.dateTime, last_c.close
        if not l2_done:
            last_c = candles[-1]
            l2_exit_time, l2_exit_price = last_c.dateTime, last_c.close

        # P&L calculations
        l1_exit_value = half_qty * (l1_exit_price or 0)
        l2_exit_value = half_qty * (l2_exit_price or 0)
        total_exit_value = round(l1_exit_value + l2_exit_value, 2)
        net_pnl = round(total_exit_value - trade_value_entry, 2)
        pnl_pct = round((net_pnl / trade_value_entry * 100), 4) if trade_value_entry else 0
        exit_price_combined = round((l1_exit_price or 0) + (l2_exit_price or 0), 4)

        # Trade summary
        if l1_exit_price == sl_value or l2_exit_price == current_sl:
            summary_text = "Trade Closed at Stop Loss"
        elif l2_exit_price == highest_t_value:
            summary_text = "Highest Target Achieved"
        elif last_hit_target_value:
            summary_text = f"Trade Closed at Trailing Stop Loss"
        else:
            summary_text = "Trade Closed at End-of-Day"

    # ====================================================================
    # SINGLE TRADE MODE
    # ====================================================================
    else:
        highest_t_value = _highest_target(signal)
        current_sl = sl_value
        l1_exit_time = l2_exit_time = None
        l1_exit_price = l2_exit_price = None
        last_hit_target_value = None

        iter_start = 1 if req.entry_time == "At Signal" else 2
        trade_candles = candles[iter_start:]

        for candle in trade_candles:
            c_date = candle.dateTime.date()
            trade_type = (signal.trade_type or "Intraday").strip()

            # Global Rule: expiry force-close
            if expiry_date and c_date > expiry_date:
                last_expiry_candle = _last_candle_of_day(candles_by_date, expiry_date)
                close_price = last_expiry_candle.close if last_expiry_candle else final_entry_price
                close_time = last_expiry_candle.dateTime if last_expiry_candle else candle.dateTime
                l1_exit_time = l2_exit_time = close_time
                l1_exit_price = l2_exit_price = close_price
                break

            # Intraday EOD close
            if trade_type == "Intraday" and req.closing_on_trade_type == "Close at EOD Values":
                if c_date > signal_date:
                    last_c = _last_candle_of_day(candles_by_date, signal_date)
                    cp = last_c.close if last_c else final_entry_price
                    ct = last_c.dateTime if last_c else candle.dateTime
                    l1_exit_time = l2_exit_time = ct
                    l1_exit_price = l2_exit_price = cp
                    break

            # BTST next-day EOD close
            if trade_type == "BTST" and req.closing_on_trade_type == "Close at EOD Values":
                next_day = signal_date + timedelta(days=1)
                if c_date > next_day:
                    last_c = _last_candle_of_day(candles_by_date, next_day)
                    cp = last_c.close if last_c else final_entry_price
                    ct = last_c.dateTime if last_c else candle.dateTime
                    l1_exit_time = l2_exit_time = ct
                    l1_exit_price = l2_exit_price = cp
                    break

            # Record target hit times
            for idx, t_val in all_targets:
                if target_hit_times[idx] is None and candle.high >= t_val:
                    target_hit_times[idx] = candle.dateTime

            # SL hit time
            if candle.low <= sl_value and sl_hit_time is None:
                sl_hit_time = candle.dateTime

            # Update trailing SL
            new_sl = current_sl
            for idx, t_val in all_targets:
                if candle.high >= t_val:
                    if idx == 1:
                        new_sl = max(new_sl, final_entry_price)
                    else:
                        prev_val = all_targets[idx - 2][1]
                        new_sl = max(new_sl, prev_val)
                    last_hit_target_value = t_val
            current_sl = new_sl

            # SL check (trailing)
            if candle.low <= current_sl:
                l1_exit_time = l2_exit_time = candle.dateTime
                l1_exit_price = l2_exit_price = current_sl
                break

            # Highest target check
            if highest_t_value and candle.high >= highest_t_value:
                l1_exit_time = l2_exit_time = candle.dateTime
                l1_exit_price = l2_exit_price = highest_t_value
                break

        # Fallback
        if l1_exit_time is None:
            last_c = candles[-1]
            l1_exit_time = l2_exit_time = last_c.dateTime
            l1_exit_price = l2_exit_price = last_c.close

        exit_price_combined = l1_exit_price or 0
        total_exit_value = round(qty * exit_price_combined, 2)
        net_pnl = round(total_exit_value - trade_value_entry, 2)
        pnl_pct = round((net_pnl / trade_value_entry * 100), 4) if trade_value_entry else 0

        if l1_exit_price == sl_value or l1_exit_price == current_sl:
            summary_text = "Trade Closed at Stop Loss"
        elif l1_exit_price == highest_t_value:
            summary_text = "Highest Target Achieved"
        elif last_hit_target_value:
            summary_text = "Trade Closed at Trailing Stop Loss"
        else:
            summary_text = "Trade Closed at End-of-Day"

    # ====================================================================
    # Build Report Rows
    # ====================================================================
    # T1..T10 VALUES for Summary Report (Report 1)
    t_values = {}
    for i in range(1, 11):
        attr = getattr(signal, f"target_{i}", None)
        t_values[f"T{i}"] = round(attr, 4) if attr else "-"

    # T1..T10 HIT TIMES for Validation Report (Report 2)
    t_times = {}
    for idx, _ in all_targets:
        hit_dt = target_hit_times.get(idx)
        t_times[f"T{idx}"] = _format_time(hit_dt)
    for i in range(1, 11):
        if f"T{i}" not in t_times:
            t_times[f"T{i}"] = "-"

    summary_row = {
        "serial": serial_num,
        "script": signal.script,
        "entry_type": entry_type_label,
        "option_type": signal.type,
        "qty": qty,
        "entry_time": _format_dt(entry_time),
        "entry_price": round(final_entry_price, 4),
        "trade_value": trade_value_entry,
        "sl": round(sl_value, 4),
        **t_values,
        "l1_exit_time": _format_dt(l1_exit_time),
        "l1_exit_price": round(l1_exit_price, 4) if l1_exit_price else None,
        "l2_exit_time": _format_dt(l2_exit_time),
        "l2_exit_price": round(l2_exit_price, 4) if l2_exit_price else None,
        "exit_price": round(exit_price_combined, 4),
        "net_pnl": net_pnl,
        "pnl_pct": pnl_pct,
    }

    validation_row = {
        "serial": serial_num,
        "script": signal.script,
        "option_type": signal.type,
        "expiry": signal.expiry,
        "qty": qty,
        "entry_time": _format_dt(entry_time),
        "entry_price": round(final_entry_price, 4),
        "trade_amount_at_entry": trade_value_entry,
        "sl": _format_time(sl_hit_time),
        **t_times,
        "l1_exit_time": _format_dt(l1_exit_time),
        "l1_exit_price": round(l1_exit_price, 4) if l1_exit_price else None,
        "l2_exit_time": _format_dt(l2_exit_time),
        "l2_exit_price": round(l2_exit_price, 4) if l2_exit_price else None,
        "trade_amount_at_exit": total_exit_value,
        "trade_summary": summary_text,
        "trade_pnl": net_pnl,
        "pnl_pct": pnl_pct,
    }

    return summary_row, validation_row


# ---------------------------------------------------------------------------
# Background Validation Job
# ---------------------------------------------------------------------------

def _run_signal_validation(job_id: str, req: SignalValidateRequest, db_url: str):
    """
    Background job: fetches signals, fetches matching OHLC candles per signal,
    runs _run_trade() for each, and persists SignalValidationReport.

    Isolated from _run_validation() in validator.py — no shared code paths.
    REUSABLE: The job pattern (status dict + background thread) mirrors validator.py.
    """
    from sqlalchemy import event
    from sqlmodel import create_engine
    engine = create_engine(db_url, connect_args={"check_same_thread": False})

    # REUSABLE: Same WAL + busy_timeout pattern as database.py main engine.
    # Prevents "database is locked" when validation runs alongside ingestion.
    @event.listens_for(engine, "connect")
    def _set_sqlite_pragma(dbapi_conn, _):
        cursor = dbapi_conn.cursor()
        cursor.execute("PRAGMA journal_mode=WAL;")
        cursor.execute("PRAGMA busy_timeout=15000;")
        cursor.close()

    try:
        with Session(engine) as session:
            _sv_jobs[job_id]["status"] = "running"

            # --- 1. Fetch Signals ---
            signal_stmt = select(SignalData).where(
                SignalData.signal_provider == req.signal_provider
            ).order_by(SignalData.dateTime)

            if req.stock and req.stock.upper() != "ALL":
                signal_stmt = signal_stmt.where(SignalData.stock == req.stock)

            if req.start_date:
                try:
                    start_dt = datetime.strptime(req.start_date, "%Y-%m-%d")
                    signal_stmt = signal_stmt.where(SignalData.dateTime >= start_dt)
                except ValueError:
                    pass
            if req.end_date:
                try:
                    end_dt = datetime.strptime(req.end_date, "%Y-%m-%d")
                    end_dt = end_dt.replace(hour=23, minute=59, second=59)
                    signal_stmt = signal_stmt.where(SignalData.dateTime <= end_dt)
                except ValueError:
                    pass

            # Filter by trade_on — DB stores "CE" / "PE" (not "Call" / "Put")
            if req.trade_on == "Calls":
                signal_stmt = signal_stmt.where(SignalData.type == "CE")
            elif req.trade_on == "Puts":
                signal_stmt = signal_stmt.where(SignalData.type == "PE")

            signals = session.exec(signal_stmt).all()

            if not signals:
                _sv_jobs[job_id] = {
                    "status": "error",
                    "result": {"error": "No signals found for the given criteria."}
                }
                return

            # --- 2. Prefetch all relevant OHLC data ---
            # Query a wide date window covering all signals + buffer for expiry
            min_dt = min(s.dateTime for s in signals)
            max_dt = max(s.dateTime for s in signals)
            # Add 2 extra days buffer for BTST and expiry close
            ohlc_end = max_dt + timedelta(days=7)

            ohlc_stmt = select(OptionsData).where(
                OptionsData.dateTime >= min_dt,
                OptionsData.dateTime <= ohlc_end,
            ).order_by(OptionsData.dateTime)

            if req.stock and req.stock.upper() != "ALL":
                ohlc_stmt = ohlc_stmt.where(OptionsData.stock == req.stock)

            all_ohlc = session.exec(ohlc_stmt).all()

            # Issue 3 fix: Index OHLC by (script, expiry, type) to keep
            # Call and Put candles separate. optionsdata.type = 'Call'/'Put'.
            ohlc_index: dict = {}
            for c in all_ohlc:
                key = (c.script, c.expiry, c.type)   # type included to avoid Call/Put mixing
                if key not in ohlc_index:
                    ohlc_index[key] = []
                ohlc_index[key].append(c)

            # Index OHLC by date for EOD lookups
            # Key: (script, expiry, type, date) → [candle, ...]
            ohlc_by_date: dict = {}
            for c in all_ohlc:
                key = (c.script, c.expiry, c.type, c.dateTime.date())
                if key not in ohlc_by_date:
                    ohlc_by_date[key] = []
                ohlc_by_date[key].append(c)

            # --- 3. Run trade engine for each signal ---
            summary_trades = []
            validation_trades = []
            serial = 1

            skipped_signals = 0  # Issue 4: track signals with no OHLC match

            for signal in signals:
                # Issue 3 fix: map signal type CE/PE → Call/Put to match optionsdata.type
                # Issue 2 fix: cast script int → str to match optionsdata.script (string)
                signal_type_str = "Call" if signal.type == "CE" else "Put"
                script_key = (str(signal.script), signal.expiry, signal_type_str)
                all_signal_candles = ohlc_index.get(script_key, [])

                # Issue 4: signals with abnormal/unmapped scripts (e.g. CRUDEOIL strikes
                # ingested under a NIFTY filter) will have no OHLC match — skip gracefully.
                if not all_signal_candles:
                    skipped_signals += 1
                    continue

                # Filter candles from signal datetime onwards
                signal_candles = [c for c in all_signal_candles if c.dateTime >= signal.dateTime]

                if not signal_candles:
                    skipped_signals += 1
                    continue

                # Build per-date index for this script+expiry+type (EOD lookups)
                per_date: dict = {}
                for c in all_signal_candles:
                    d = c.dateTime.date()
                    if d not in per_date:
                        per_date[d] = []
                    per_date[d].append(c)

                s_row, v_row = _run_trade(signal, signal_candles, per_date, req, serial)
                if s_row is not None:
                    summary_trades.append(s_row)
                    validation_trades.append(v_row)
                    serial += 1

            if not summary_trades:
                _sv_jobs[job_id] = {
                    "status": "error",
                    "result": {"error": "No trades could be executed. Check that Options OHLC data exists for the signal scripts and date range."}
                }
                return

            # --- 4. Compute summary stats ---
            total_trades = len(summary_trades)
            profitable = sum(1 for r in summary_trades if r["net_pnl"] > 0)
            win_rate = round((profitable / total_trades * 100), 2) if total_trades else 0
            total_pnl = round(sum(r["net_pnl"] for r in summary_trades), 2)

            # --- 5. Persist report ---
            report = SignalValidationReport(
                config=json.dumps(req.dict()),
                signal_provider=req.signal_provider,
                stock=req.stock,
                lot_split_rule=req.lot_split_rule,
                total_trades=total_trades,
                total_pnl=total_pnl,
                win_rate=win_rate,
                summary_trades=json.dumps(summary_trades),
                validation_trades=json.dumps(validation_trades),
            )
            session.add(report)
            session.commit()
            session.refresh(report)

            _sv_jobs[job_id] = {
                "status": "done",
                "result": {
                    "reportId": report.id,
                    "signal_provider": req.signal_provider,
                    "stock": req.stock,
                    "lot_split_rule": req.lot_split_rule,
                    "total_trades": total_trades,
                    "win_rate": win_rate,
                    "total_pnl": total_pnl,
                    "skipped_signals": skipped_signals,   # Issue 4: signals with no OHLC match
                    "summary_trades": summary_trades,
                    "validation_trades": validation_trades,
                }
            }

    except Exception as e:
        _sv_jobs[job_id] = {
            "status": "error",
            "result": {"error": str(e)}
        }


# ---------------------------------------------------------------------------
# API Endpoints
# ---------------------------------------------------------------------------

@router.post("")
async def start_signal_validation(
    req: SignalValidateRequest,
    background_tasks: BackgroundTasks,
):
    """
    Starts a signal validation job in the background.
    Returns a jobId for polling.
    Pattern mirrors /api/validate in validator.py — completely isolated.
    """
    job_id = str(uuid.uuid4())
    _sv_jobs[job_id] = {"status": "pending"}
    background_tasks.add_task(_run_signal_validation, job_id, req, sqlite_url)
    return {"jobId": job_id}


@router.get("/status/{job_id}")
def get_signal_validation_status(job_id: str):
    """
    Returns the current status and result of a signal validation job.
    Mirrors /api/validate/status/{jobId} pattern from validator.py.
    """
    job = _sv_jobs.get(job_id)
    if not job:
        return {"status": "not_found"}
    return {"status": job["status"], "result": job.get("result")}


@router.get("/reports")
def list_signal_reports(session: Session = Depends(get_session)):
    """
    Returns all past signal validation reports (summary only, no trade detail).
    Used to populate the signal reports dropdown in the Results tab.
    """
    stmt = select(SignalValidationReport).order_by(SignalValidationReport.testDate.desc())
    reports = session.exec(stmt).all()
    return [
        {
            "id": r.id,
            "testDate": r.testDate.isoformat() if r.testDate else None,
            "signal_provider": r.signal_provider,
            "stock": r.stock,
            "lot_split_rule": r.lot_split_rule,
            "total_trades": r.total_trades,
            "total_pnl": r.total_pnl,
            "win_rate": r.win_rate,
        }
        for r in reports
    ]


@router.get("/reports/{report_id}")
def get_signal_report(report_id: str, session: Session = Depends(get_session)):
    """
    Returns a specific signal validation report with all trade data.
    """
    report = session.get(SignalValidationReport, report_id)
    if not report:
        return {"error": "Report not found"}
    return {
        "id": report.id,
        "testDate": report.testDate.isoformat() if report.testDate else None,
        "signal_provider": report.signal_provider,
        "stock": report.stock,
        "lot_split_rule": report.lot_split_rule,
        "total_trades": report.total_trades,
        "total_pnl": report.total_pnl,
        "win_rate": report.win_rate,
        "config": json.loads(report.config) if report.config else {},
        "summary_trades": json.loads(report.summary_trades) if report.summary_trades else [],
        "validation_trades": json.loads(report.validation_trades) if report.validation_trades else [],
    }


@router.get("/export-excel")
def export_signal_report_excel(
    resultId: str,
    reportType: str = "summary",  # "summary" | "validation"
    session: Session = Depends(get_session)
):
    """
    Downloads Signal Validation Report as Excel (.xlsx).
    reportType="summary" → Report 1 format (target values)
    reportType="validation" → Report 2 format (target hit times)
    """
    from fastapi.responses import StreamingResponse
    import io

    report = session.get(SignalValidationReport, resultId)
    if not report:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Report not found")

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl is required for Excel export. Install it via pip.")

    wb = openpyxl.Workbook()
    ws = wb.active

    if reportType == "validation":
        trades = json.loads(report.validation_trades) if report.validation_trades else []
        ws.title = "Signal Validation Report"
        headers = [
            "#", "Script / Strike Price", "Option Type", "Expiry", "Qty",
            "Entry Time", "Entry Price", "Trade Amount at Entry",
            "SL (Time Hit)", "T1", "T2", "T3", "T4", "T5", "T6", "T7", "T8", "T9", "T10",
            "L1-Exit Time", "L1-Exit Price", "L2-Exit Time", "L2-Exit Price",
            "Trade Amount at Exit", "Trade Summary", "Trade Profit/Loss", "PnL %"
        ]
        row_keys = [
            "serial", "script", "option_type", "expiry", "qty",
            "entry_time", "entry_price", "trade_amount_at_entry",
            "sl", "T1", "T2", "T3", "T4", "T5", "T6", "T7", "T8", "T9", "T10",
            "l1_exit_time", "l1_exit_price", "l2_exit_time", "l2_exit_price",
            "trade_amount_at_exit", "trade_summary", "trade_pnl", "pnl_pct"
        ]
    else:
        trades = json.loads(report.summary_trades) if report.summary_trades else []
        ws.title = "Summary Report"
        headers = [
            "#", "Script / Strike Price", "Entry Type", "Option Type", "Qty",
            "Entry Time", "Entry Price", "Trade Value",
            "SL", "T1", "T2", "T3", "T4", "T5", "T6", "T7", "T8", "T9", "T10",
            "L1-Exit Time", "L1-Exit Price", "L2-Exit Time", "L2-Exit Price",
            "Exit Price", "Net P&L", "P&L %"
        ]
        row_keys = [
            "serial", "script", "entry_type", "option_type", "qty",
            "entry_time", "entry_price", "trade_value",
            "sl", "T1", "T2", "T3", "T4", "T5", "T6", "T7", "T8", "T9", "T10",
            "l1_exit_time", "l1_exit_price", "l2_exit_time", "l2_exit_price",
            "exit_price", "net_pnl", "pnl_pct"
        ]

    # Header styling
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(color="000000", bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    profit_font = Font(color="4EDEA3", bold=True)
    loss_font = Font(color="FF4444", bold=True)
    for trade in trades:
        row = [trade.get(k, "") for k in row_keys]
        ws.append(row)
        # Colour Net P&L cell
        pnl_idx = row_keys.index("net_pnl") if "net_pnl" in row_keys else row_keys.index("trade_pnl") if "trade_pnl" in row_keys else None
        if pnl_idx is not None:
            cell = ws.cell(row=ws.max_row, column=pnl_idx + 1)
            pnl_val = trade.get("net_pnl") or trade.get("trade_pnl") or 0
            cell.font = profit_font if (pnl_val or 0) >= 0 else loss_font
        # Append "%" symbol to PnL % cell
        if "pnl_pct" in row_keys:
            pct_cell = ws.cell(row=ws.max_row, column=row_keys.index("pnl_pct") + 1)
            if pct_cell.value != "" and pct_cell.value is not None:
                pct_cell.value = f"{pct_cell.value}%"

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"signal_report_{reportType}_{resultId[:8]}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
